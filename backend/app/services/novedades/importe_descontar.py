"""Importación masiva 'Importe a descontar' para Capital Humano."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from uuid import uuid4

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.novedades import NovedadesAjusteCapital, NovedadesPeriodo, PeriodoEstado
from app.schemas.novedades import (
    ImporteDescontarAnularResponse,
    ImporteDescontarImportResponse,
    ImporteDescontarStatusResponse,
    ModuloImportRowError,
)
from app.services.novedades.capital_humano import build_capital_humano_rows
from app.services.novedades.export_xls import build_grid_rows
from app.services.novedades.helpers import soft_delete

REQUIRED_HEADERS = ("Legajo", "Nombre y Apellido", "Sector", "Monto")


@dataclass
class _ParsedRow:
    row_num: int
    legajo: str
    nombre: str
    sector: str
    monto_abs: Decimal
    professional_id: int


def _raise_errors(message: str, errors: list[ModuloImportRowError]) -> None:
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail={"message": message, "errors": [{"row": e.row, "reason": e.reason} for e in errors]},
    )


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_monto(raw: str) -> Decimal | None:
    if raw is None or str(raw).strip() == "":
        return None
    text = str(raw).strip().replace(" ", "").replace("$", "")
    if "," in text and "." in text:
        # 1.234,56 → EU
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _require_closed_periodo(db: Session, periodo_id: int) -> NovedadesPeriodo:
    periodo = db.execute(
        select(NovedadesPeriodo).where(
            NovedadesPeriodo.id == periodo_id,
            NovedadesPeriodo.deleted_at.is_(None),
        )
    ).scalar_one_or_none()
    if not periodo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Periodo no encontrado")
    if periodo.estado != PeriodoEstado.closed:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Importe a descontar solo disponible para períodos cerrados",
        )
    return periodo


def _active_lote_id(db: Session, periodo_id: int) -> str | None:
    item = db.execute(
        select(NovedadesAjusteCapital.descuento_lote_id)
        .where(
            NovedadesAjusteCapital.periodo_id == periodo_id,
            NovedadesAjusteCapital.deleted_at.is_(None),
            NovedadesAjusteCapital.descuento_lote_id.is_not(None),
        )
        .limit(1)
    ).scalar_one_or_none()
    return item


def status_importe_descontar(db: Session, periodo_id: int) -> ImporteDescontarStatusResponse:
    _require_closed_periodo(db, periodo_id)
    lote = _active_lote_id(db, periodo_id)
    return ImporteDescontarStatusResponse(has_descuento=bool(lote), lote_id=lote)


def anular_importe_descontar(db: Session, periodo_id: int, actor_id: int) -> ImporteDescontarAnularResponse:
    _require_closed_periodo(db, periodo_id)
    items = list(
        db.execute(
            select(NovedadesAjusteCapital).where(
                NovedadesAjusteCapital.periodo_id == periodo_id,
                NovedadesAjusteCapital.deleted_at.is_(None),
                NovedadesAjusteCapital.descuento_lote_id.is_not(None),
            )
        )
        .scalars()
        .all()
    )
    for item in items:
        soft_delete(item, actor_id)
    db.commit()
    return ImporteDescontarAnularResponse(deleted=len(items))


def _cargas_by_servicio(db: Session, periodo_id: int) -> dict[int, list[tuple[int, Decimal]]]:
    """professional_id -> [(servicio_id, monto_cargas)] sorted by monto desc, sid asc."""
    detail = build_grid_rows(db, periodo_id=periodo_id, servicio_id=None, q=None, concepto_q=None)
    totals: dict[int, dict[int, Decimal]] = {}
    for row in detail:
        pid = row.professional_id
        sid = row.servicio_id
        if sid is None:
            continue
        totals.setdefault(pid, {})
        totals[pid][sid] = totals[pid].get(sid, Decimal("0")) + Decimal(row.valor or 0)

    out: dict[int, list[tuple[int, Decimal]]] = {}
    for pid, by_svc in totals.items():
        ordered = sorted(by_svc.items(), key=lambda x: (-x[1], x[0]))
        out[pid] = ordered
    return out


def _waterfill(
    services: list[tuple[int, Decimal]], abs_discount: Decimal
) -> list[tuple[int | None, Decimal]]:
    """Return [(servicio_id|None, negative_importe), ...]."""
    if not services:
        return [(None, -abs_discount)]

    remaining = abs_discount
    allocations: list[tuple[int | None, Decimal]] = []
    for idx, (sid, capacidad) in enumerate(services):
        is_last = idx == len(services) - 1
        if is_last:
            take = remaining
        else:
            take = min(remaining, capacidad) if capacidad > 0 else Decimal("0")
            if take <= 0 and remaining > 0 and idx < len(services) - 1:
                continue
        if take <= 0 and not is_last:
            continue
        if take > 0 or (is_last and remaining > 0):
            if is_last:
                take = remaining
            allocations.append((sid, -take))
            remaining -= take
        if remaining <= 0 and not is_last:
            break
    if remaining > 0:
        # Should not happen if caller capped; dump on last service
        if allocations:
            sid, prev = allocations[-1]
            allocations[-1] = (sid, prev - remaining)
        else:
            allocations.append((services[-1][0], -remaining))
    return [(sid, amt) for sid, amt in allocations if amt != 0]


def import_importe_descontar(
    db: Session, *, periodo_id: int, content: bytes, actor_id: int
) -> ImporteDescontarImportResponse:
    _require_closed_periodo(db, periodo_id)
    if _active_lote_id(db, periodo_id):
        _raise_errors(
            "Ya hay un descuento importado",
            [ModuloImportRowError(row=0, reason="Anulá el descuento actual antes de importar otro archivo")],
        )

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        _raise_errors(
            "Archivo Excel inválido",
            [ModuloImportRowError(row=0, reason="No se pudo leer el archivo")],
        )
        raise exc  # pragma: no cover

    ws = wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        _raise_errors(
            "Archivo vacío",
            [ModuloImportRowError(row=1, reason="Falta fila de encabezados")],
        )

    headers = [_cell_str(h) for h in header_row]
    # Exact match required (presence + text); order may vary
    col_index: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h in REQUIRED_HEADERS:
            col_index[h] = idx
    missing = [name for name in REQUIRED_HEADERS if name not in col_index]
    if missing:
        _raise_errors(
            "Encabezados incompletos",
            [ModuloImportRowError(row=1, reason=f"Faltan o no coinciden columnas: {', '.join(missing)}")],
        )

    grid_rows = build_capital_humano_rows(db, periodo_id=periodo_id, include_bonos=True)
    by_legajo: dict[str, object] = {}
    for row in grid_rows:
        key = (row.legajo or "").strip()
        if key:
            by_legajo[key] = row
            by_legajo[key.casefold()] = row

    cargas_map = _cargas_by_servicio(db, periodo_id)

    errors: list[ModuloImportRowError] = []
    seen_legajos: set[str] = set()
    parsed: list[_ParsedRow] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or _cell_str(v) == "" for v in row):
            continue

        def col(name: str) -> str:
            idx = col_index[name]
            if idx >= len(row):
                return ""
            return _cell_str(row[idx])

        legajo = col("Legajo")
        nombre = col("Nombre y Apellido")
        sector = col("Sector")
        monto_raw = col("Monto")

        if not legajo and not nombre and not sector and not monto_raw:
            continue

        row_errors: list[str] = []
        if not legajo:
            row_errors.append("legajo obligatorio")
        else:
            leg_key = legajo.casefold()
            if leg_key in seen_legajos:
                row_errors.append(f"legajo duplicado en el archivo: {legajo}")
            else:
                seen_legajos.add(leg_key)

        monto_val = _parse_monto(monto_raw)
        if monto_val is None:
            row_errors.append("monto inválido o vacío")
        elif monto_val == 0:
            row_errors.append("monto no puede ser 0")

        grid_row = None
        if legajo:
            grid_row = by_legajo.get(legajo) or by_legajo.get(legajo.casefold())
            if not grid_row:
                row_errors.append(f"legajo no está en Capital Humano del período: {legajo}")

        if row_errors:
            errors.append(ModuloImportRowError(row=row_num, reason="; ".join(row_errors)))
            continue

        assert grid_row is not None and monto_val is not None
        abs_discount = abs(monto_val)
        cargas = Decimal(grid_row.monto_cargas or 0)
        produccion = Decimal(grid_row.monto_bonos or 0)
        ajustes_exist = Decimal(grid_row.monto_ajustes or 0)
        cap = cargas + produccion
        projected = cargas + ajustes_exist + produccion - abs_discount

        if abs_discount > cap:
            errors.append(
                ModuloImportRowError(
                    row=row_num,
                    reason=(
                        f"legajo {legajo}: descuento {abs_discount} supera cargas+producción ({cap})"
                    ),
                )
            )
            continue
        if projected < 0:
            errors.append(
                ModuloImportRowError(
                    row=row_num,
                    reason=f"legajo {legajo}: el total general quedaría negativo ({projected})",
                )
            )
            continue

        parsed.append(
            _ParsedRow(
                row_num=row_num,
                legajo=legajo,
                nombre=nombre,
                sector=sector,
                monto_abs=abs_discount,
                professional_id=grid_row.professional_id,
            )
        )

    if errors:
        _raise_errors("No se importó ningún descuento", errors)

    if not parsed:
        _raise_errors(
            "Archivo sin filas",
            [ModuloImportRowError(row=0, reason="No hay filas de datos para importar")],
        )

    lote_id = str(uuid4())
    now = datetime.utcnow()
    created = 0

    for item in parsed:
        neg = -item.monto_abs
        comentario = f"{item.legajo} - {item.nombre} - {item.sector} - {neg}"[:500]
        services = cargas_map.get(item.professional_id) or []
        allocations = _waterfill(services, item.monto_abs)
        for sid, importe in allocations:
            db.add(
                NovedadesAjusteCapital(
                    professional_id=item.professional_id,
                    periodo_id=periodo_id,
                    servicio_id=sid,
                    importe=importe,
                    comentario=comentario,
                    descuento_lote_id=lote_id,
                    created_at=now,
                    updated_at=now,
                    created_by=actor_id,
                    updated_by=actor_id,
                    deleted_at=None,
                )
            )
            created += 1

    db.commit()
    return ImporteDescontarImportResponse(created=created, lote_id=lote_id)
