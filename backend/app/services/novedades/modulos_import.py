"""Bulk Excel import for Novedades módulos (Param)."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.novedades import NovedadesModulo, NovedadesServicio
from app.schemas.novedades import ModuloCreateRequest, ModuloImportResponse, ModuloImportRowError
from app.services.novedades import masters as masters_service

HEADERS = ("descripcion", "comentario", "valor", "produccion", "sadofe", "servicio")
LIST_SHEET = "_servicios"


@dataclass
class _ParsedRow:
    row_num: int
    descripcion: str
    comentario: str | None
    valor: Decimal
    produccion: bool
    sadofe: bool
    servicio_id: int


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_si_no(raw: str, *, field: str) -> bool | str:
    """Return bool or error reason string."""
    text = (raw or "").strip().casefold()
    if text in ("sí", "si", "s", "yes", "y", "1", "true"):
        return True
    if text in ("no", "n", "0", "false"):
        return False
    if not text:
        return False
    return f"{field} debe ser Sí o No"


def _parse_valor(raw: str) -> Decimal | str:
    text = (raw or "").strip()
    if not text:
        return Decimal("0")
    text = text.replace(",", ".")
    try:
        value = Decimal(text)
    except (InvalidOperation, ValueError):
        return "valor inválido"
    if value < 0:
        return "valor debe ser ≥ 0"
    return value


def build_modulos_import_template(db: Session) -> bytes:
    servicios = masters_service.list_servicios(db, only_active=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "modulos"
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 18 if header != "comentario" else 28

    list_ws = wb.create_sheet(LIST_SHEET)
    for idx, servicio in enumerate(servicios, start=1):
        list_ws.cell(row=idx, column=1, value=servicio.nombre)
    list_ws.sheet_state = "hidden"

    max_row = max(len(servicios), 1)
    servicio_dv = DataValidation(
        type="list",
        formula1=f"'{LIST_SHEET}'!$A$1:$A${max_row}",
        allow_blank=True,
        showDropDown=False,
    )
    servicio_dv.error = "Elegí un servicio de la lista"
    servicio_dv.errorTitle = "Servicio"
    ws.add_data_validation(servicio_dv)
    servicio_dv.add("F2:F5001")

    si_no_dv = DataValidation(
        type="list",
        formula1='"Sí,No"',
        allow_blank=True,
        showDropDown=False,
    )
    si_no_dv.error = "Usá Sí o No"
    si_no_dv.errorTitle = "Valor"
    ws.add_data_validation(si_no_dv)
    si_no_dv.add("D2:E5001")

    # Hint row (optional example style — leave blank for user fill)
    ws.cell(row=2, column=4, value="No")
    ws.cell(row=2, column=5, value="No")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _servicio_map(db: Session) -> dict[str, NovedadesServicio]:
    rows = masters_service.list_servicios(db, only_active=True)
    return {s.nombre.strip().casefold(): s for s in rows}


def _existing_descripciones(db: Session) -> set[str]:
    rows = db.execute(
        select(NovedadesModulo.descripcion).where(NovedadesModulo.deleted_at.is_(None))
    ).scalars().all()
    return {str(d).strip().casefold() for d in rows if d}


def _normalize_header(value) -> str:
    text = _cell_str(value).casefold()
    text = text.replace("ó", "o").replace("í", "i")
    return re.sub(r"\s+", "_", text)


def import_modulos_from_xlsx(db: Session, content: bytes, actor_id: int) -> ModuloImportResponse:
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"message": "Archivo Excel inválido", "errors": [{"row": 0, "reason": "No se pudo leer el archivo"}]},
        ) from exc

    ws = wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"message": "Archivo vacío", "errors": [{"row": 1, "reason": "Falta fila de encabezados"}]},
        )

    headers = [_normalize_header(h) for h in header_row]
    col_index: dict[str, int] = {}
    aliases = {
        "descripcion": "descripcion",
        "descripción": "descripcion",
        "comentario": "comentario",
        "valor": "valor",
        "produccion": "produccion",
        "producción": "produccion",
        "sadofe": "sadofe",
        "servicio": "servicio",
    }
    for idx, h in enumerate(headers):
        key = aliases.get(h)
        if key:
            col_index[key] = idx

    missing = [name for name in ("descripcion", "servicio") if name not in col_index]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "Encabezados incompletos",
                "errors": [{"row": 1, "reason": f"Faltan columnas: {', '.join(missing)}"}],
            },
        )

    servicios_by_name = _servicio_map(db)
    existing_desc = _existing_descripciones(db)
    seen_in_file: set[str] = set()
    errors: list[ModuloImportRowError] = []
    parsed: list[_ParsedRow] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or _cell_str(v) == "" for v in row):
            continue

        def col(name: str) -> str:
            idx = col_index.get(name)
            if idx is None or idx >= len(row):
                return ""
            return _cell_str(row[idx])

        descripcion = col("descripcion")
        servicio_nombre = col("servicio")
        comentario_raw = col("comentario") if "comentario" in col_index else ""
        valor_raw = col("valor") if "valor" in col_index else ""
        produccion_raw = col("produccion") if "produccion" in col_index else "No"
        sadofe_raw = col("sadofe") if "sadofe" in col_index else "No"

        # Skip completely empty hint row leftovers
        if not descripcion and not servicio_nombre and not valor_raw and not comentario_raw:
            continue

        row_errors: list[str] = []
        if not descripcion:
            row_errors.append("descripción obligatoria")
        elif len(descripcion) < 2:
            row_errors.append("descripción demasiado corta")
        elif len(descripcion) > 200:
            row_errors.append("descripción demasiado larga")

        desc_key = descripcion.casefold()
        if descripcion:
            if desc_key in existing_desc:
                row_errors.append("ya existe un módulo con esa descripción")
            if desc_key in seen_in_file:
                row_errors.append("descripción duplicada en el archivo")
            else:
                seen_in_file.add(desc_key)

        servicio_id = 0
        if not servicio_nombre:
            row_errors.append("servicio obligatorio")
        else:
            servicio = servicios_by_name.get(servicio_nombre.casefold())
            if not servicio:
                row_errors.append(f"servicio no encontrado o inactivo: {servicio_nombre}")
            else:
                servicio_id = servicio.id

        valor = _parse_valor(valor_raw)
        if isinstance(valor, str):
            row_errors.append(valor)

        produccion = _parse_si_no(produccion_raw, field="produccion")
        if isinstance(produccion, str):
            row_errors.append(produccion)

        sadofe = _parse_si_no(sadofe_raw, field="sadofe")
        if isinstance(sadofe, str):
            row_errors.append(sadofe)

        if row_errors:
            errors.append(ModuloImportRowError(row=row_num, reason="; ".join(row_errors)))
            continue

        parsed.append(
            _ParsedRow(
                row_num=row_num,
                descripcion=descripcion,
                comentario=comentario_raw or None,
                valor=valor,  # type: ignore[arg-type]
                produccion=bool(produccion),
                sadofe=bool(sadofe),
                servicio_id=servicio_id,
            )
        )

    if errors:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "No se importó ningún módulo. Corregí los errores e intentá de nuevo.",
                "errors": [e.model_dump() for e in errors],
            },
        )

    if not parsed:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "No hay filas para importar",
                "errors": [{"row": 0, "reason": "El archivo no tiene datos"}],
            },
        )

    for item in parsed:
        masters_service.create_modulo(
            db,
            ModuloCreateRequest(
                descripcion=item.descripcion,
                comentario=item.comentario,
                valor=item.valor,
                produccion=item.produccion,
                sadofe=item.sadofe,
                servicio_ids=[item.servicio_id],
            ),
            actor_id=actor_id,
            commit=False,
        )
    db.commit()
    return ModuloImportResponse(created=len(parsed), errors=[])
