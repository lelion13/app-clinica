from io import BytesIO
from decimal import Decimal
from unittest.mock import MagicMock

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.schemas.novedades import ModuloCreateRequest
from app.services.novedades import modulos_import as mi


def _xlsx_bytes(rows: list[tuple]):
    wb = Workbook()
    ws = wb.active
    ws.append(["descripcion", "comentario", "valor", "produccion", "sadofe", "servicio"])
    for row in rows:
        ws.append(list(row))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_si_no():
    assert mi._parse_si_no("Sí", field="produccion") is True
    assert mi._parse_si_no("no", field="produccion") is False
    assert mi._parse_si_no("", field="produccion") is False
    assert isinstance(mi._parse_si_no("maybe", field="produccion"), str)


def test_parse_valor_empty_is_zero():
    assert mi._parse_valor("") == Decimal("0")
    assert mi._parse_valor("10,5") == Decimal("10.5")
    assert isinstance(mi._parse_valor("-1"), str)


def test_template_has_headers_and_validation(monkeypatch):
    db = MagicMock()
    servicio = MagicMock()
    servicio.nombre = "Guardia"
    monkeypatch.setattr(mi.masters_service, "list_servicios", lambda _db, only_active=False: [servicio])
    content = mi.build_modulos_import_template(db)
    wb = load_workbook(BytesIO(content))
    ws = wb["modulos"]
    headers = [c.value for c in ws[1]]
    assert headers == ["descripcion", "comentario", "valor", "produccion", "sadofe", "servicio"]
    assert "_servicios" in wb.sheetnames


def test_import_all_or_nothing_on_duplicate(monkeypatch):
    db = MagicMock()
    servicio = MagicMock()
    servicio.id = 7
    servicio.nombre = "Guardia"

    monkeypatch.setattr(mi.masters_service, "list_servicios", lambda _db, only_active=False: [servicio])
    monkeypatch.setattr(mi, "_existing_descripciones", lambda _db: {"existente"})

    created = []

    def fake_create(db, payload, actor_id, *, commit=True):
        created.append(payload)
        return MagicMock()

    monkeypatch.setattr(mi.masters_service, "create_modulo", fake_create)

    content = _xlsx_bytes(
        [
            ("Nuevo OK", "", "100", "No", "No", "Guardia"),
            ("Existente", "", "50", "Sí", "No", "Guardia"),
        ]
    )
    with pytest.raises(HTTPException) as exc:
        mi.import_modulos_from_xlsx(db, content, actor_id=1)
    assert exc.value.status_code == 400
    assert created == []
    errors = exc.value.detail["errors"]
    assert any(e["row"] == 3 for e in errors)


def test_import_ok_commits(monkeypatch):
    db = MagicMock()
    servicio = MagicMock()
    servicio.id = 3
    servicio.nombre = "UTI"

    monkeypatch.setattr(mi.masters_service, "list_servicios", lambda _db, only_active=False: [servicio])
    monkeypatch.setattr(mi, "_existing_descripciones", lambda _db: set())

    created = []

    def fake_create(db, payload: ModuloCreateRequest, actor_id, *, commit=True):
        assert commit is False
        created.append(payload)
        return MagicMock()

    monkeypatch.setattr(mi.masters_service, "create_modulo", fake_create)

    content = _xlsx_bytes(
        [
            ("Mod A", "c1", "", "Sí", "No", "UTI"),
            ("Mod B", "", "20", "No", "Sí", "uti"),
        ]
    )
    result = mi.import_modulos_from_xlsx(db, content, actor_id=9)
    assert result.created == 2
    assert created[0].valor == Decimal("0")
    assert created[0].produccion is True
    assert created[0].servicio_ids == [3]
    assert created[1].sadofe is True
    db.commit.assert_called_once()


def test_import_unknown_servicio(monkeypatch):
    db = MagicMock()
    monkeypatch.setattr(mi.masters_service, "list_servicios", lambda _db, only_active=False: [])
    monkeypatch.setattr(mi, "_existing_descripciones", lambda _db: set())
    content = _xlsx_bytes([("Mod", "", "1", "No", "No", "Inexistente")])
    with pytest.raises(HTTPException) as exc:
        mi.import_modulos_from_xlsx(db, content, actor_id=1)
    assert "servicio no encontrado" in exc.value.detail["errors"][0]["reason"]
